
import os
import csv

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QListWidget,
    QListWidgetItem, QMessageBox, QFileDialog, QFormLayout, QComboBox, QCheckBox, QFrame
)
from PySide6.QtCore import Qt

from core.config import COLORS, EXPORT_DIR
from core.database import db
from ui.widgets.shamsi_date_picker import ShamsiDateEdit

TABLES = {
    "personnel": "پرسنل",
    "incidents": "حوادث و شبه‌حوادث",
    "ppe_items": "انبار PPE",
    "ppe_issuance": "تحویل PPE",
    "training_courses": "دوره‌های آموزشی",
    "training_records": "ثبت دوره‌های آموزشی",
    "medical_exams": "معاینات طب کار",
    "disciplinary_records": "تشویق و تنبیه",
}

# Mapping table to its date column and personnel_id column
TABLE_META = {
    "personnel": {"date_col": "hire_date_shamsi", "personnel_col": "id"},
    "incidents": {"date_col": "incident_date_shamsi", "personnel_col": "personnel_id"},
    "ppe_items": {"date_col": "created_at", "personnel_col": None},
    "ppe_issuance": {"date_col": "issue_date_shamsi", "personnel_col": "personnel_id"},
    "training_courses": {"date_col": "created_at", "personnel_col": None},
    "training_records": {"date_col": "completion_date_shamsi", "personnel_col": "personnel_id"},
    "medical_exams": {"date_col": "exam_date_shamsi", "personnel_col": "personnel_id"},
    "disciplinary_records": {"date_col": "event_date_shamsi", "personnel_col": "personnel_id"},
}

class ReportsPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)

        header = QLabel("گزارش‌گیری و خروجی داده‌ها")
        header.setStyleSheet(f"color: {COLORS['text_primary']}; font-size: 18px; font-weight: bold;")
        root.addWidget(header)

        # --- Filters Section ---
        filter_frame = QFrame()
        filter_frame.setStyleSheet(f"background-color: {COLORS['bg_card']}; border-radius: 8px;")
        filter_layout = QFormLayout(filter_frame)
        filter_layout.setContentsMargins(15, 15, 15, 15)

        # Checkbox to enable filtering
        self.enable_filters_cb = QCheckBox("اعمال فیلتر (بازه زمانی و پرسنل)")
        self.enable_filters_cb.setStyleSheet(f"color: {COLORS['text_primary']};")
        self.enable_filters_cb.toggled.connect(self._toggle_filters)
        filter_layout.addRow(self.enable_filters_cb)

        # Date pickers
        date_layout = QHBoxLayout()
        self.from_date = ShamsiDateEdit()
        self.to_date = ShamsiDateEdit()

        lbl_from = QLabel("از تاریخ:")
        lbl_from.setStyleSheet(f"color: {COLORS['text_secondary']};")
        lbl_to = QLabel("تا تاریخ:")
        lbl_to.setStyleSheet(f"color: {COLORS['text_secondary']};")

        date_layout.addWidget(lbl_from)
        date_layout.addWidget(self.from_date)
        date_layout.addWidget(lbl_to)
        date_layout.addWidget(self.to_date)
        filter_layout.addRow("بازه زمانی:", date_layout)

        # Personnel Multiple Selection (ListWidget instead of ComboBox for multiple select)
        self.personnel_list = QListWidget()
        self.personnel_list.setSelectionMode(QListWidget.MultiSelection)
        self.personnel_list.setFixedHeight(120)
        self.personnel_list.setStyleSheet(f"""
            QListWidget {{
                background-color: {COLORS['bg_input']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
            }}
            QListWidget::item:selected {{
                background-color: {COLORS['accent']};
                color: #1B2129;
            }}
        """)
        filter_layout.addRow("انتخاب پرسنل:", self.personnel_list)

        root.addWidget(filter_frame)

        # Disable filters initially
        self._toggle_filters(False)

        # --- Tables Section ---
        desc = QLabel("یک یا چند جدول را انتخاب کرده و فرمت خروجی را مشخص کنید.")
        desc.setStyleSheet(f"color: {COLORS['text_secondary']};")
        root.addWidget(desc)

        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QListWidget.MultiSelection)
        self.list_widget.setStyleSheet(f"""
            QListWidget {{
                background-color: {COLORS['bg_card']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                padding: 6px;
            }}
            QListWidget::item:selected {{
                background-color: {COLORS['accent']};
                color: #1B2129;
            }}
        """)
        for key, label in TABLES.items():
            item = QListWidgetItem(label)
            item.setData(1000, key)
            self.list_widget.addItem(item)
        root.addWidget(self.list_widget, 1)

        btn_row = QHBoxLayout()
        excel_btn = QPushButton("خروجی Excel (.xlsx)")
        csv_btn = QPushButton("خروجی CSV")
        for b, color in [(excel_btn, COLORS["accent"]), (csv_btn, COLORS["safety_blue"])]:
            b.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    color: #1B2129;
                    border: none;
                    border-radius: 6px;
                    padding: 10px 16px;
                    font-weight: bold;
                }}
            """)
        excel_btn.clicked.connect(self._export_excel)
        csv_btn.clicked.connect(self._export_csv)
        btn_row.addWidget(excel_btn)
        btn_row.addWidget(csv_btn)
        root.addLayout(btn_row)

    def _toggle_filters(self, state):
        self.from_date.setEnabled(state)
        self.to_date.setEnabled(state)
        self.personnel_list.setEnabled(state)

    def reload(self):
        self.personnel_list.clear()
        rows = db.fetch_all("SELECT id, personnel_code, first_name, last_name FROM personnel ORDER BY last_name, first_name")
        for r in rows:
            item = QListWidgetItem(f"{r['first_name']} {r['last_name']} ({r['personnel_code']})")
            item.setData(Qt.UserRole, r["id"])
            self.personnel_list.addItem(item)

    def _selected_tables(self):
        return [item.data(1000) for item in self.list_widget.selectedItems()]

    def _build_query(self, table_name):
        query = f"SELECT * FROM {table_name}"
        params = []
        conditions = []

        if self.enable_filters_cb.isChecked():
            meta = TABLE_META.get(table_name)
            if meta:
                # Date filtering
                from_d = self.from_date.value()
                to_d = self.to_date.value()
                date_col = meta["date_col"]

                if from_d:
                    conditions.append(f"{date_col} >= ?")
                    params.append(from_d)
                if to_d:
                    conditions.append(f"{date_col} <= ?")
                    params.append(to_d)

                # Personnel filtering
                personnel_col = meta["personnel_col"]
                selected_p_items = self.personnel_list.selectedItems()
                if personnel_col and selected_p_items:
                    p_ids = [str(item.data(Qt.UserRole)) for item in selected_p_items]
                    placeholders = ",".join(["?"] * len(p_ids))
                    conditions.append(f"{personnel_col} IN ({placeholders})")
                    params.extend(p_ids)

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        return query, tuple(params)

    def _export_csv(self):
        tables = self._selected_tables()
        if not tables:
            QMessageBox.warning(self, "خطا", "حداقل یک جدول را انتخاب کنید.")
            return
        folder = QFileDialog.getExistingDirectory(self, "انتخاب پوشه ذخیره‌سازی", EXPORT_DIR)
        if not folder:
            return
        for t in tables:
            query, params = self._build_query(t)
            rows = db.fetch_all(query, params)
            path = os.path.join(folder, f"{t}.csv")
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                if rows:
                    writer.writerow(rows[0].keys())
                    for r in rows:
                        writer.writerow(list(r))
                else:
                    writer.writerow(["(بدون داده)"])
        QMessageBox.information(self, "موفق", f"فایل‌های CSV در مسیر زیر ذخیره شدند:\n{folder}")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            QMessageBox.critical(
                self, "خطا",
                "کتابخانه openpyxl نصب نیست. دستور نصب:\npip install openpyxl"
            )
            return

        tables = self._selected_tables()
        if not tables:
            QMessageBox.warning(self, "خطا", "حداقل یک جدول را انتخاب کنید.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره فایل اکسل", os.path.join(EXPORT_DIR, "hse_report.xlsx"), "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for t in tables:
            query, params = self._build_query(t)
            rows = db.fetch_all(query, params)
            ws = wb.create_sheet(title=TABLES.get(t, t)[:31])
            if rows:
                ws.append(list(rows[0].keys()))
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                for r in rows:
                    ws.append(list(r))
            else:
                ws.append(["(بدون داده)"])
        wb.save(path)
        QMessageBox.information(self, "موفق", f"فایل اکسل ذخیره شد:\n{path}")
